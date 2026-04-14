"""
import_data.py — parse and import GlobalData Excel exports.

Accepts one or more .xlsx files, finds the header row automatically,
maps columns to the database schema, detects FID and contractor signals
from project descriptions, and upserts projects + contacts into Supabase.

Sub-projects (project type containing "sub") are skipped and counted
separately so the user knows how many were filtered out.
"""

import io
import re
import logging
from datetime import datetime, date as date_type
from typing import List

import openpyxl
from fastapi import APIRouter, Depends, UploadFile, File

from app.config import (
    TABLE_PROJECTS,
    TABLE_CONTACTS,
    HEADER_SEARCH_MAX_ROWS,
    HEADER_DATA_OFFSET,
    CONTACT_SLOTS,
    IMPORT_CHUNK_SIZE,
)
from app.lib.auth import get_current_user
from app.lib.clients import get_admin_client
from app.lib.categorize import auto_categorize, normalize_stage
from app.lib.regions import get_world_region
from app.lib.detect import detect_fid, detect_contractor
from app.lib.utils import chunk
from app.schemas import ImportResponse

logger = logging.getLogger(__name__)
router  = APIRouter()


# ---------------------------------------------------------------------------
# Column name aliases
# Keys are internal field names; values are accepted header spellings
# (lowercase, compared after collapsing whitespace).
# ---------------------------------------------------------------------------
COLUMN_MAP = {
    "globaldata_id":     ["project id", "projectid", "id"],
    "name":              ["project name", "projectname", "name"],
    "company_name":      ["company name", "company", "client", "owner"],
    "country":           ["country"],
    "status":            ["project status", "status"],
    "project_value_usd": ["project value (usd)", "project value", "value (usd)", "value"],
    "execution_date":    ["execution date", "start date", "completion date"],
    "description":       ["project description", "description", "summary"],
    "project_url":       ["project url", "url", "link", "globaldata url"],
    "sector":            ["primary sector", "sector"],
    "momentum_score":    ["momentum score", "momentum"],
    "project_type":      ["project type", "type"],
    "city":              ["city"],
    "region":            ["region", "state", "province"],
}


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

def find_col(headers: list[str], keys: list[str]) -> int | None:
    """Return the index of the first header that matches any of the given keys."""
    for i, h in enumerate(headers):
        clean = re.sub(r"\s+", " ", str(h).lower().strip())
        for key in keys:
            if key == clean or key in clean:
                return i
    return None


def safe_str(val) -> str | None:
    """Convert a cell value to a stripped string, or None for blanks/NaN."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s.lower() not in ["none", "nan"] else None


def parse_value(raw) -> int | None:
    """Parse a numeric cell (possibly with currency symbols / commas) to int."""
    if raw is None:
        return None
    try:
        return int(float(str(raw).replace(",", "").replace("$", "").strip()))
    except (ValueError, TypeError):
        return None


def parse_date(raw) -> str | None:
    """
    Parse a cell value to an ISO-format date string.
    openpyxl returns datetime/date objects for proper date cells;
    string cells are tried against common formats.
    """
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date().isoformat()
    if isinstance(raw, date_type):
        return raw.isoformat()
    s = str(raw).strip()
    if not s or s.lower() in ["none", "nan", ""]:
        return None
    for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d",
                "%d-%m-%Y", "%b %Y", "%B %Y"]:
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return s  # return as-is; scoring will handle unparseable dates gracefully


def parse_contacts(row: list, headers: list[str]) -> list[dict]:
    """
    Extract up to CONTACT_SLOTS key contacts from a project row.
    Skips slots where no name is found.
    """
    contacts = []
    for i in range(1, CONTACT_SLOTS + 1):
        name_idx  = find_col(headers, [f"key person name {i}", f"key_person_name_{i}"])
        title_idx = find_col(headers, [f"key contact {i}", f"key_contact_{i}", f"contact title {i}"])
        email_idx = find_col(headers, [f"email {i}", f"contact email {i}", f"key email {i}"])
        lin_idx   = find_col(headers, [f"linkedin {i}", f"linkedin url {i}"])

        name = safe_str(row[name_idx] if name_idx is not None and name_idx < len(row) else None)
        if not name:
            continue  # no name = skip this slot

        contacts.append({
            "name":         name,
            "title":        safe_str(row[title_idx]) if title_idx is not None and title_idx < len(row) else "",
            "email":        safe_str(row[email_idx]) if email_idx is not None and email_idx < len(row) else None,
            "linkedin_url": safe_str(row[lin_idx])   if lin_idx  is not None and lin_idx  < len(row) else None,
        })
    return contacts


# ---------------------------------------------------------------------------
# Import endpoint
# ---------------------------------------------------------------------------

@router.post("/import", response_model=ImportResponse)
async def import_projects(
    files: List[UploadFile] = File(...),
    user=Depends(get_current_user),
):
    """
    Import one or more GlobalData Excel files (.xlsx).

    For each file:
    - Locates the header row (searches the first HEADER_SEARCH_MAX_ROWS rows)
    - Skips the units/blank rows after the header (HEADER_DATA_OFFSET)
    - Parses each project row and upserts it into the database
    - Extracts contacts and refreshes them per project

    Returns total imported, sub-projects skipped, and any per-file errors.
    """
    supabase       = get_admin_client()
    total_imported = 0
    total_subs     = 0
    errors: list[str] = []

    for file in files:
        if not file.filename or not file.filename.lower().endswith(".xlsx"):
            errors.append(f"{file.filename}: not an xlsx file — skipped")
            continue

        content = await file.read()
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
        except Exception as e:
            errors.append(f"{file.filename}: could not open file — {e}")
            continue

        # Locate the header row by scanning the first N rows
        headers        = []
        header_row_idx = None
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=HEADER_SEARCH_MAX_ROWS, values_only=True)
        ):
            row_vals = [str(c).lower().strip() if c is not None else "" for c in row]
            if any("project" in v or "name" in v or "country" in v for v in row_vals):
                headers        = row_vals
                header_row_idx = row_idx + 1  # 1-based row number
                break

        if not headers or header_row_idx is None:
            errors.append(f"{file.filename}: could not find header row in first {HEADER_SEARCH_MAX_ROWS} rows")
            continue

        col_map = {field: find_col(headers, keys) for field, keys in COLUMN_MAP.items()}

        projects_to_upsert: list[dict] = []
        contacts_to_insert: list[dict] = []

        # HEADER_DATA_OFFSET rows after the header before data begins
        # (GlobalData exports: row after header = column units, then one blank row)
        data_start_row = header_row_idx + HEADER_DATA_OFFSET + 1

        for row in ws.iter_rows(min_row=data_start_row, values_only=True):
            if all(c is None for c in row):
                continue  # skip fully empty rows

            def get(field: str) -> str | None:
                idx = col_map.get(field)
                return safe_str(row[idx]) if idx is not None and idx < len(row) else None

            def get_raw(field: str):
                idx = col_map.get(field)
                return row[idx] if idx is not None and idx < len(row) else None

            # Skip sub-projects (they inflate the list without adding value)
            ptype = get("project_type") or ""
            if "sub" in ptype.lower():
                total_subs += 1
                continue

            raw_id  = get("globaldata_id")
            name    = get("name") or "Unnamed project"
            desc    = get("description")
            country = get("country")
            status  = get("status") or ""
            sector  = get("sector")

            fid                              = detect_fid(desc)
            contractor_detected, contractor_name = detect_contractor(desc)
            project_contacts                 = parse_contacts(list(row), headers)

            project = {
                "globaldata_id":       raw_id or f"gd-imp-{total_imported}",
                "user_id":             user.id,
                "name":                name,
                "company_name":        get("company_name"),
                "country":             country,
                "region":              " ".join(filter(None, [get("city"), get("region")])) or None,
                "world_region":        get_world_region(country),
                "sector":              auto_categorize(name, desc, sector),
                "stage_normalized":    normalize_stage(status),
                "status":              status,
                "project_value_usd":   parse_value(get_raw("project_value_usd")),
                "execution_date":      parse_date(get_raw("execution_date")),
                "description":         desc,
                "project_url":         get("project_url"),
                "momentum_score":      parse_value(get_raw("momentum_score")),
                "fid_detected":        fid,
                "contractor_detected": contractor_detected,
                "contractor_name":     contractor_name,
                "key_contacts":        project_contacts,
            }
            projects_to_upsert.append(project)

            for c in project_contacts:
                contacts_to_insert.append({**c, "globaldata_id": project["globaldata_id"]})

        if projects_to_upsert:
            for batch in chunk(projects_to_upsert, IMPORT_CHUNK_SIZE):
                supabase.table(TABLE_PROJECTS).upsert(
                    batch, on_conflict="globaldata_id,user_id"
                ).execute()

            if contacts_to_insert:
                # Resolve globaldata_id → internal project UUID
                gd_ids    = [p["globaldata_id"] for p in projects_to_upsert]
                proj_rows = supabase.table(TABLE_PROJECTS).select(
                    "id, globaldata_id"
                ).in_("globaldata_id", gd_ids).eq("user_id", user.id).execute()
                gd_to_id  = {r["globaldata_id"]: r["id"] for r in proj_rows.data or []}

                final_contacts = []
                for c in contacts_to_insert:
                    gd_id = c.pop("globaldata_id")
                    pid   = gd_to_id.get(gd_id)
                    if pid:
                        final_contacts.append({
                            "project_id":   pid,
                            "name":         c["name"],
                            "title":        c.get("title") or "",
                            "email":        c.get("email"),
                            "linkedin_url": c.get("linkedin_url"),
                            "source":       "globaldata",
                            "role_type":    "other",
                        })

                if final_contacts:
                    # Replace old GlobalData contacts with freshly imported ones
                    project_ids = list(gd_to_id.values())
                    for id_batch in chunk(project_ids, IMPORT_CHUNK_SIZE):
                        supabase.table(TABLE_CONTACTS).delete().in_(
                            "project_id", id_batch
                        ).eq("source", "globaldata").execute()
                    for batch in chunk(final_contacts, IMPORT_CHUNK_SIZE):
                        supabase.table(TABLE_CONTACTS).insert(batch).execute()

        total_imported += len(projects_to_upsert)
        logger.info(
            "Imported %d projects from %s (%d sub-projects skipped).",
            len(projects_to_upsert), file.filename, total_subs,
        )

    return {
        "ok":                   True,
        "imported":             total_imported,
        "sub_projects_removed": total_subs,
        "errors":               errors,
    }
